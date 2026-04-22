import os
import re
import json
import time
import hashlib
from io import BytesIO
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor, as_completed

import pandas as pd
import requests
from flask import Flask, jsonify, render_template, request, send_file, session
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

"""
Public-safe demo version of a multi-warehouse dispatch analytics tool.

What was sanitized:
- Removed internal endpoint details and replaced them with env vars / placeholders
- Removed hardcoded secret values
- Generalized warehouse metadata
- Added clear config guards so the app can be shared publicly without exposing credentials

Before running, set:
  APP_SECRET_KEY
  API_BASE_URL
  API_LOGIN_PATH
  API_HISTORY_PATH
  API_REALTIME_PATH

Optional:
  WAREHOUSE_CONFIG_PATH
"""

app = Flask(__name__)
app.secret_key = os.getenv("APP_SECRET_KEY", "change-me-in-production")

app.config["SESSION_COOKIE_SAMESITE"] = "Lax"
app.config["SESSION_COOKIE_SECURE"] = False
app.config["SESSION_COOKIE_HTTPONLY"] = True

API_BASE_URL = os.getenv("API_BASE_URL", "https://api.example.com")
API_LOGIN_PATH = os.getenv("API_LOGIN_PATH", "/auth/login")
API_HISTORY_PATH = os.getenv("API_HISTORY_PATH", "/dispatch/history")
API_REALTIME_PATH = os.getenv("API_REALTIME_PATH", "/dispatch/realtime")

DEFAULT_WAREHOUSES = {
    "WH1": {"name": "Warehouse 1", "branch_id": 101, "batch_prefix": "WH1SUB"},
    "WH2": {"name": "Warehouse 2", "branch_id": 102, "batch_prefix": "WH2SUB"},
    "WH3": {"name": "Warehouse 3", "branch_id": 103, "batch_prefix": "WH3SUB"},
}

CONFIG_PATH = os.getenv("WAREHOUSE_CONFIG_PATH")
if CONFIG_PATH and os.path.exists(CONFIG_PATH):
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        MAIN_WAREHOUSES = json.load(f)
else:
    MAIN_WAREHOUSES = DEFAULT_WAREHOUSES

BRANCH_TO_MAIN = {v["branch_id"]: k for k, v in MAIN_WAREHOUSES.items()}

QUERY_CACHE = {}
QUERY_CACHE_TTL_SECONDS = 600

MAX_WAREHOUSE_WORKERS = 4
MAX_BATCH_WORKERS = 16


def _ensure_configured() -> None:
    required = {
        "APP_SECRET_KEY": app.secret_key,
        "API_BASE_URL": API_BASE_URL,
        "API_LOGIN_PATH": API_LOGIN_PATH,
        "API_HISTORY_PATH": API_HISTORY_PATH,
        "API_REALTIME_PATH": API_REALTIME_PATH,
    }
    missing = [k for k, v in required.items() if not v]
    if missing:
        raise RuntimeError(f"Missing required configuration: {', '.join(missing)}")


def normalize_team_name(name):
    if not name:
        return "Unknown"
    return str(name).strip()


def build_cache_key(payload):
    raw = json.dumps(payload, sort_keys=True, ensure_ascii=False)
    return hashlib.md5(raw.encode("utf-8")).hexdigest()


def cache_get(key):
    item = QUERY_CACHE.get(key)
    if not item:
        return None
    if time.time() - item["ts"] > QUERY_CACHE_TTL_SECONDS:
        QUERY_CACHE.pop(key, None)
        return None
    return item["data"]


def cache_set(key, data):
    QUERY_CACHE[key] = {
        "ts": time.time(),
        "data": data,
    }


def extract_batches_from_dispatch_details(dispatch_details):
    if not dispatch_details:
        return []
    parts = [x.strip() for x in str(dispatch_details).split(",") if x.strip()]
    seen = set()
    out = []
    for p in parts:
        if p not in seen:
            seen.add(p)
            out.append(p)
    return out


def looks_like_delivery_record(name):
    if not name:
        return False
    s = str(name).strip().upper()
    return bool(re.search(r"DELI", s))


def parse_mmdd_from_name(name):
    if not name:
        return None

    s = str(name).strip().upper()

    m1 = re.search(r"\b(\d{4})\b", s)
    if m1:
        mm = int(m1.group(1)[:2])
        dd = int(m1.group(1)[2:])
        if 1 <= mm <= 12 and 1 <= dd <= 31:
            return mm, dd

    m2 = re.search(r"(\d{1,2})/(\d{1,2})", s)
    if m2:
        mm = int(m2.group(1))
        dd = int(m2.group(2))
        if 1 <= mm <= 12 and 1 <= dd <= 31:
            return mm, dd

    return None


def record_matches_business_window(record_name, start_date_str, end_date_str):
    mmdd = parse_mmdd_from_name(record_name)
    if not mmdd:
        return False

    start_dt = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    end_dt = datetime.strptime(end_date_str, "%Y-%m-%d").date()

    year = start_dt.year
    try:
        named_day = datetime(year, mmdd[0], mmdd[1]).date()
    except ValueError:
        return False

    return start_dt <= named_day <= end_dt


def create_time_in_window(create_time_value, start_date_str, end_date_str):
    if not create_time_value:
        return False

    start_dt = datetime.strptime(start_date_str, "%Y-%m-%d").date()
    end_dt = datetime.strptime(end_date_str, "%Y-%m-%d").date()

    try:
        row_day = datetime.fromtimestamp(int(create_time_value)).date()
    except Exception:
        return False

    return start_dt <= row_day <= end_dt


def is_excluded_all_batches_name(name):
    if not name:
        return False

    raw = str(name)
    s = raw.strip().upper()

    if "PICKUP" in s:
        return True
    if "TRUCKING" in s:
        return True
    if re.search(r"(^|[\s\-_\/])WR($|[\s\-_\/])", s):
        return True

    return False


@app.route("/")
def index():
    wh_list = [{"code": k, "name": v["name"]} for k, v in MAIN_WAREHOUSES.items()]
    return render_template("index.html", warehouses=wh_list)


@app.route("/api/login", methods=["POST"])
def login():
    _ensure_configured()

    payload = request.json or {}
    try:
        response = requests.post(
            f"{API_BASE_URL}{API_LOGIN_PATH}",
            json=payload,
            timeout=30,
        )
        response.raise_for_status()
        data = response.json()
    except Exception as e:
        return jsonify(success=False, message=f"Login request failed: {e}")

    token = ((data.get("data") or {}).get("token")) or data.get("token")
    if not token:
        return jsonify(success=False, message="Login failed")

    session["token"] = token
    return jsonify(success=True)


def _fetch_dispatch_history(branch_id, token):
    try:
        response = requests.get(
            f"{API_BASE_URL}{API_HISTORY_PATH}",
            headers={"Authorization": f"Bearer {token}"},
            params={"branch": branch_id, "page": 1, "offset": 10000},
            timeout=60,
        )
        response.raise_for_status()
        data = response.json()
        rows = (((data.get("data") or {}).get("data")) or [])
        return branch_id, rows, None
    except Exception as e:
        return branch_id, [], str(e)


def _fetch_batch_realtime(sub_reference, token):
    try:
        response = requests.get(
            f"{API_BASE_URL}{API_REALTIME_PATH}",
            headers={"Authorization": f"Bearer {token}"},
            params={"sub_references": sub_reference, "team_id": 0},
            timeout=45,
        )
        response.raise_for_status()
        data = response.json().get("data") or {}
        rows = data.get("delivery_data") or []
        total_count = int(data.get("total_count") or 0)
        undo_count = int(data.get("undo_count") or 0)
        total_done_ratio = data.get("total_done_ratio") or ""

        return sub_reference, {
            "delivery_data": rows,
            "total_count": total_count,
            "undo_count": undo_count,
            "total_done_ratio": total_done_ratio,
        }, None
    except Exception as e:
        return sub_reference, None, str(e)


def collect_candidate_batches(main_code, start_date_str, end_date_str, history_rows, mode):
    wh = MAIN_WAREHOUSES[main_code]
    prefix = wh["batch_prefix"] + "-"
    batches = set()

    for row in history_rows:
        if int(row.get("is_removed") or 0) != 0:
            continue

        name = row.get("name", "")
        dispatch_details = row.get("dispatch_details", "")

        if not dispatch_details:
            continue

        if mode == "delivery_203":
            if not looks_like_delivery_record(name):
                continue
            if not record_matches_business_window(name, start_date_str, end_date_str):
                continue

        elif mode in ("all_valid_203", "all_valid_total"):
            if not create_time_in_window(row.get("create_time"), start_date_str, end_date_str):
                continue
            if is_excluded_all_batches_name(name):
                continue
        else:
            continue

        for batch in extract_batches_from_dispatch_details(dispatch_details):
            if batch.startswith(prefix):
                batches.add(batch)

    return sorted(batches)


def get_metric_label(mode):
    if mode in ("delivery_203", "all_valid_203"):
        return "203"
    if mode == "all_valid_total":
        return "Total"
    return "Count"


def fetch_data(payload):
    selected = payload.get("warehouses", [])
    start_date = payload.get("start_date")
    end_date = payload.get("end_date")
    mode = payload.get("mode", "delivery_203")
    token = session.get("token", "")

    if not selected:
        raise ValueError("No warehouses selected")
    if not start_date or not end_date:
        raise ValueError("Missing date range")
    if not token:
        raise ValueError("Please sign in first")
    if mode not in ("delivery_203", "all_valid_203", "all_valid_total"):
        mode = "delivery_203"

    cache_payload = {
        "selected": selected,
        "start_date": start_date,
        "end_date": end_date,
        "mode": mode,
    }
    cache_key = build_cache_key(cache_payload)
    cached = cache_get(cache_key)
    if cached is not None:
        return cached

    selected_branch_ids = [
        MAIN_WAREHOUSES[code]["branch_id"]
        for code in selected
        if code in MAIN_WAREHOUSES
    ]

    history_by_main = {}
    history_logs = []

    with ThreadPoolExecutor(max_workers=min(MAX_WAREHOUSE_WORKERS, max(1, len(selected_branch_ids)))) as ex:
        futures = {
            ex.submit(_fetch_dispatch_history, branch_id, token): branch_id
            for branch_id in selected_branch_ids
        }
        for fut in as_completed(futures):
            branch_id, rows, err = fut.result()
            main_code = BRANCH_TO_MAIN.get(branch_id, "")
            if main_code:
                history_by_main[main_code] = rows

            history_logs.append({
                "warehouse_id": main_code or str(branch_id),
                "history_count": len(rows),
                **({"error": err} if err else {}),
            })

    batches_by_main = {}
    all_batches = set()

    for main_code in selected:
        rows = history_by_main.get(main_code, [])
        candidate_batches = collect_candidate_batches(main_code, start_date, end_date, rows, mode)
        batches_by_main[main_code] = candidate_batches
        for batch in candidate_batches:
            all_batches.add(batch)

    batch_result_map = {}
    batch_fetch_errors = []

    all_batches = sorted(all_batches)

    with ThreadPoolExecutor(max_workers=min(MAX_BATCH_WORKERS, max(1, len(all_batches)))) as ex:
        futures = {
            ex.submit(_fetch_batch_realtime, batch, token): batch
            for batch in all_batches
        }
        for fut in as_completed(futures):
            batch, data, err = fut.result()
            if data is not None:
                batch_result_map[batch] = data
            else:
                batch_fetch_errors.append({"batch": batch, "error": err})

    dsp_agg = {}
    main_agg = {}
    driver_agg = {}

    used_batches = 0
    matched_total_packages = 0
    global_driver_ids = set()

    for main_code in selected:
        main_agg.setdefault(main_code, {
            "metric_count": 0,
            "matched_total_packages": 0,
            "driver_ids": set(),
            "dsp_keys": set(),
            "batch_count": 0,
        })

        for batch in batches_by_main.get(main_code, []):
            batch_data = batch_result_map.get(batch)
            if not batch_data:
                continue

            used_batches += 1
            main_agg[main_code]["batch_count"] += 1

            batch_total_count = int(batch_data.get("total_count") or 0)
            matched_total_packages += batch_total_count
            main_agg[main_code]["matched_total_packages"] += batch_total_count

            rows = batch_data.get("delivery_data", []) or []

            for row in rows:
                team_id = row.get("team_id", "")
                team_name = normalize_team_name(row.get("team_name", "Unknown"))
                driver_id = row.get("shipping_staff_id", "")

                metric_value = int(row.get("total_count") or 0) if mode == "all_valid_total" else int(row.get("203") or 0)

                key = (main_code, str(team_id), team_name)

                if key not in dsp_agg:
                    dsp_agg[key] = {
                        "Main Warehouse": main_code,
                        "DSP ID": str(team_id) if team_id != "" else "",
                        "DSP Name": team_name,
                        "Package_Count": 0,
                        "Driver_IDs": set(),
                    }

                dsp_agg[key]["Package_Count"] += metric_value

                if driver_id not in ("", None):
                    driver_str = str(driver_id)
                    dsp_agg[key]["Driver_IDs"].add(driver_str)
                    main_agg[main_code]["driver_ids"].add(driver_str)
                    global_driver_ids.add(driver_str)

                main_agg[main_code]["metric_count"] += metric_value
                main_agg[main_code]["dsp_keys"].add(key)

                if driver_id not in ("", None):
                    driver_key = str(driver_id)
                    if driver_key not in driver_agg:
                        driver_agg[driver_key] = {
                            "Driver ID": str(driver_id),
                            "Driver Name": str(driver_id),
                            "Package Count": 0,
                        }
                    driver_agg[driver_key]["Package Count"] += metric_value

    main_summary_rows = []
    for main_code in selected:
        summary = main_agg.get(main_code, {
            "metric_count": 0,
            "driver_ids": set(),
            "dsp_keys": set(),
        })
        main_summary_rows.append({
            "Main Warehouse": main_code,
            "Package_Count": int(summary["metric_count"]),
            "Driver_Count": len(summary["driver_ids"]),
            "DSP_Count": len(summary["dsp_keys"]),
        })

    main_summary_df = pd.DataFrame(main_summary_rows)
    if main_summary_df.empty:
        main_summary_df = pd.DataFrame(columns=["Main Warehouse", "Package_Count", "Driver_Count", "DSP_Count"])

    dsp_summary_rows = []
    for row in dsp_agg.values():
        dsp_summary_rows.append({
            "Main Warehouse": row["Main Warehouse"],
            "DSP ID": row["DSP ID"],
            "DSP Name": row["DSP Name"],
            "Package_Count": int(row["Package_Count"]),
            "Driver_Count": len(row["Driver_IDs"]),
        })

    dsp_summary_df = pd.DataFrame(dsp_summary_rows)
    if dsp_summary_df.empty:
        dsp_summary_df = pd.DataFrame(columns=["Main Warehouse", "DSP ID", "DSP Name", "Package_Count", "Driver_Count"])
    else:
        dsp_summary_df = dsp_summary_df.sort_values(
            ["Main Warehouse", "Package_Count", "DSP Name"],
            ascending=[True, False, True],
        )

    driver_chart_df = pd.DataFrame(list(driver_agg.values()))
    if driver_chart_df.empty:
        driver_chart_df = pd.DataFrame(columns=["Driver ID", "Driver Name", "Package Count"])
    else:
        driver_chart_df = driver_chart_df.sort_values(["Package Count"], ascending=[False]).head(20).fillna("")

    warehouse_chart_df = main_summary_df[["Main Warehouse", "Package_Count"]].rename(
        columns={"Package_Count": "Package Count"}
    )

    metric_total = int(main_summary_df["Package_Count"].sum()) if not main_summary_df.empty else 0
    active_drivers = len(global_driver_ids)
    active_dsps = int(len(dsp_summary_df)) if not dsp_summary_df.empty else 0

    top_wh = ""
    if not main_summary_df.empty and metric_total > 0:
        top_wh = main_summary_df.sort_values(["Package_Count"], ascending=[False]).iloc[0]["Main Warehouse"]

    metric_label = get_metric_label(mode)

    kpis = {
        "total_packages": metric_total,
        "active_drivers": active_drivers,
        "active_dsps": active_dsps,
        "top_warehouse": top_wh,
        "metric_label": metric_label,
    }

    ui_logs = []
    for main_code in selected:
        summary = main_agg.get(main_code, {"metric_count": 0, "batch_count": 0})
        ui_logs.append({
            "warehouse_id": main_code,
            "count": int(summary["metric_count"]),
            "batch_count": int(summary["batch_count"]),
        })

    result = {
        "main_summary": main_summary_df.fillna("").to_dict("records"),
        "dsp_summary": dsp_summary_df.fillna("").to_dict("records"),
        "driver_chart": driver_chart_df.fillna("").to_dict("records"),
        "warehouse_chart": warehouse_chart_df.fillna("").to_dict("records"),
        "total": metric_total,
        "logs": ui_logs,
        "kpis": kpis,
        "debug": {
            "mode": mode,
            "metric_label": metric_label,
            "history_logs": history_logs,
            "batch_fetch_errors": batch_fetch_errors,
            "batches_by_main": batches_by_main,
            "used_batches": used_batches,
            "matched_total_packages": matched_total_packages,
            "metric_total": metric_total,
        },
    }

    cache_set(cache_key, result)
    return result


@app.route("/api/query", methods=["POST"])
def query():
    if not session.get("token"):
        return jsonify(success=False, message="Please sign in first"), 401
    try:
        data = fetch_data(request.json or {})
        payload = {
            "main_summary": data["main_summary"],
            "dsp_summary": data["dsp_summary"],
            "driver_chart": data["driver_chart"],
            "warehouse_chart": data["warehouse_chart"],
            "kpis": data["kpis"],
            "total": data["total"],
            "logs": data["logs"],
            "debug": data["debug"],
        }
        return jsonify(success=True, data=payload)
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500


def style_summary_sheet(ws, df, title):
    parchment = "F2E8D0"
    dark = "2C1A06"
    gold = "B8960A"
    mid = "D4B870"
    light = "F8F2E0"
    alt = "EDE0C0"

    thin = Side(style="thin", color=mid)
    thick = Side(style="medium", color=gold)
    cell_border = Border(left=thin, right=thin, top=thin, bottom=thin)

    ws.insert_rows(1)
    ws.insert_rows(1)

    title_cell = ws.cell(row=1, column=1, value=title)
    title_cell.font = Font(name="Garamond", bold=True, size=13, color=dark)
    title_cell.fill = PatternFill("solid", fgColor=parchment)
    title_cell.alignment = Alignment(horizontal="left", vertical="center", indent=1)
    title_cell.border = Border(bottom=thick)
    ws.row_dimensions[1].height = 28
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(df.columns)))

    header_row = 3
    for ci, col in enumerate(df.columns, 1):
        cell = ws.cell(row=header_row, column=ci, value=col)
        cell.fill = PatternFill("solid", fgColor=dark)
        cell.font = Font(name="Garamond", bold=True, size=10, color=mid)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = Border(left=thick, right=thick, top=thick, bottom=thick)
    ws.row_dimensions[header_row].height = 20

    start_data = header_row + 1
    for ri, row in enumerate(df.itertuples(index=False), start_data):
        bg = alt if (ri - header_row) % 2 == 0 else light
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font = Font(name="Garamond", size=10, color=dark)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = cell_border

    for ci, col in enumerate(df.columns, 1):
        values = [str(col)]
        for row_i in range(start_data, start_data + len(df)):
            values.append(str(ws.cell(row=row_i, column=ci).value or ""))
        width = min(max(len(v) for v in values) + 4, 45)
        ws.column_dimensions[get_column_letter(ci)].width = width


@app.route("/api/export", methods=["POST"])
def export():
    if not session.get("token"):
        return jsonify(success=False, message="Please sign in first"), 401

    try:
        data = fetch_data(request.json or {})
    except Exception as e:
        return jsonify(success=False, message=str(e)), 500

    export_df = pd.DataFrame(data["dsp_summary"])
    if export_df.empty:
        export_df = pd.DataFrame(columns=["Main Warehouse", "DSP ID", "DSP Name", "Package_Count", "Driver_Count"])

    output = BytesIO()
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        export_df.to_excel(writer, sheet_name="Summary", index=False)
        ws = writer.sheets["Summary"]
        style_summary_sheet(ws, export_df, "Dispatch Analytics Summary")

    output.seek(0)
    fname = f"dispatch_analytics_summary_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(
        output,
        download_name=fname,
        as_attachment=True,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=int(os.getenv("PORT", "8088")), debug=False, use_reloader=False)
